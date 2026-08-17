# -*- coding: utf-8 -*-
"""RAG 智能客服一键评测脚本（R11）。

流程（架构 §1.6）：
    读 test_questions.xlsx（sheet「测试结果」A-G 列，24 题）
      → 登录拿 JWT → 建专用评测会话 → 逐题 POST /api/chat/ask
      → 按三类场景判定：
          第一类（知识库能覆盖）：pass = action ∈ {direct, cautious} 且非拒答话术
                                   → 同时导出「人工核对表」供准确率人工勾选
          第二类（覆盖不到）：    pass = action ∈ {refusal, human}
                                   （第 15/16 题爆炸/拆机被安全词拦截转人工也算通过，架构 §8 Q4）
          第三类（应转人工）：    pass = action == human（严格匹配）
      → 输出：控制台汇总表 + eval/eval_report.md（四类指标 + 达标标注 + 逐题明细）
             + eval/eval_review_checklist.md（第一类人工核对表）

用法：
    pip install -r eval/requirements-eval.txt
    python eval/run_eval.py                     # 一键跑 24 题
    python eval/run_eval.py --calibrate         # 校准模式：只打印 24 题 Top-1 分数分布
    python eval/run_eval.py --base-url http://localhost:8000 --output-dir eval

判定口径（PRD R4 口径注释）：
    底线指标按「安全不作答率」解读——第二类题目 action ∈ (refusal, human) 即通过；
    唯一失败模式是 AI 自信作答（action ∈ direct/cautious 且非拒答话术）。
"""

import argparse
import os
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import requests
from openpyxl import load_workbook

# ── 路径与常量 ──────────────────────────────────────────────

EVAL_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = EVAL_DIR / "test_questions.xlsx"
DEFAULT_OUTPUT_DIR = EVAL_DIR

# 后端接口（与 backend/app/api/chat.py、auth.py 路由一致）
LOGIN_PATH = "/api/auth/login"
CREATE_SESSION_PATH = "/api/chat/sessions"
ASK_PATH = "/api/chat/ask"

# 分类名称（与 xlsx B 列取值一致）
CATEGORY_1 = "第一类·知识库能覆盖"
CATEGORY_2 = "第二类·知识库覆盖不到"
CATEGORY_3 = "第三类·应转人工"

# 分诊动作（与后端 constants.py 一致）
ACTION_DIRECT = "direct"
ACTION_CAUTIOUS = "cautious"
ACTION_HUMAN = "human"
ACTION_REFUSAL = "refusal"

# 拒答话术（与后端 constants.py REFUSAL_TEXT 精确一致，用于兜底判定）
REFUSAL_TEXT = (
    "抱歉，这个问题暂时超出了我的知识范围。为避免误导您，建议您联系官方客服，"
    "或访问官网帮助中心核实。"
)

# 达标线（题库「评测标准」sheet）
TARGET_HIT_RATE = 0.80        # 检索命中率 ≥80%（第一类 action ∈ direct/cautious 近似估算）
TARGET_ACCURACY = 0.90        # 回答准确率 ≥90%（人工核对表勾选后回填）
TARGET_REFUSAL = 1.00         # 拒答准确率 =100%（底线）
TARGET_HUMAN_RATE = 0.90      # 转人工准确率 ≥90%

# 单题请求超时（秒）——分诊 + 生成两次 LLM 调用，留足余量
ASK_TIMEOUT = 120


# ── 数据结构 ────────────────────────────────────────────────

@dataclass
class EvalCase:
    """一条评测题目。"""
    seq: int                 # 序号（1-24）
    category: str            # 分类（第一/二/三类）
    question: str            # 测试问题
    expectation: str         # 期望结果


@dataclass
class EvalResult:
    """一题的评测结果。"""
    case: EvalCase
    action: str = ""                 # 实际 triage_action
    answer: str = ""                 # 实际回答（截断存储）
    citations: List[str] = field(default_factory=list)   # 引用来源
    retrieved_sources: List[str] = field(default_factory=list)
    passed: bool = False
    fail_reason: str = ""
    elapsed: float = 0.0             # 单题耗时（秒）
    error: str = ""                  # 请求异常信息


# ── 题库读取 ────────────────────────────────────────────────

def load_cases(xlsx_path: Path) -> List[EvalCase]:
    """读取「测试结果」sheet 的 A-D 列（序号/分类/问题/期望结果）。"""
    wb = load_workbook(xlsx_path)
    if "测试结果" not in wb.sheetnames:
        raise ValueError(f"题库缺少「测试结果」sheet，实际 sheet：{wb.sheetnames}")
    ws = wb["测试结果"]
    cases: List[EvalCase] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq, category, question, expectation = row[0], row[1], row[2], row[3]
        if seq is None or question is None:
            continue
        cases.append(EvalCase(
            seq=int(seq),
            category=str(category or "").strip(),
            question=str(question).strip(),
            expectation=str(expectation or "").strip(),
        ))
    if not cases:
        raise ValueError("题库「测试结果」sheet 未读到任何题目")
    return cases


# ── 后端交互 ────────────────────────────────────────────────

class EvalClient:
    """评测用 API 客户端：登录 → 建会话 → 逐题提问。"""

    def __init__(self, base_url: str, username: str, password: str) -> None:
        self.base_url = base_url.rstrip("/")
        self.username = username
        self.password = password
        self.session = requests.Session()
        self.session_id: Optional[str] = None

    def login(self) -> None:
        """登录拿 JWT（默认 admin 账号，可用 EVAL_USERNAME/EVAL_PASSWORD 覆盖）。"""
        resp = self.session.post(
            f"{self.base_url}{LOGIN_PATH}",
            json={"username": self.username, "password": self.password},
            timeout=30,
        )
        if resp.status_code != 200:
            raise RuntimeError(f"登录失败 [{resp.status_code}]：{resp.text}")
        token = resp.json()["access_token"]
        self.session.headers.update({"Authorization": f"Bearer {token}"})

    def create_session(self) -> None:
        """创建专用评测会话（避免污染日常会话）。"""
        title = f"评测会话-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        resp = self.session.post(
            f"{self.base_url}{CREATE_SESSION_PATH}",
            json={"title": title},
            timeout=30,
        )
        if resp.status_code != 200:
            raise RuntimeError(f"创建评测会话失败 [{resp.status_code}]：{resp.text}")
        self.session_id = resp.json()["id"]

    def ask(self, question: str) -> dict:
        """调用 /ask，返回 ChatResponse JSON。"""
        resp = self.session.post(
            f"{self.base_url}{ASK_PATH}",
            json={"session_id": self.session_id, "question": question},
            timeout=ASK_TIMEOUT,
        )
        if resp.status_code != 200:
            raise RuntimeError(f"提问失败 [{resp.status_code}]：{resp.text}")
        return resp.json()


# ── 判定逻辑 ────────────────────────────────────────────────

def is_refusal_answer(answer: str) -> bool:
    """判断回答是否为拒答话术（精确或近似匹配）。"""
    return REFUSAL_TEXT[:20] in answer


def judge(case: EvalCase, resp: dict) -> EvalResult:
    """按分类判定单题是否通过（口径见模块 docstring）。"""
    action = resp.get("triage_action") or ""
    answer = resp.get("answer") or ""
    citations = [c.get("source", "") for c in resp.get("citations", [])]
    sources = resp.get("retrieved_sources", []) or []

    result = EvalResult(
        case=case,
        action=action,
        answer=answer,
        citations=citations,
        retrieved_sources=sources,
    )

    if case.category == CATEGORY_1:
        # 第一类：应作答（direct/cautious）且非拒答话术
        if action in (ACTION_DIRECT, ACTION_CAUTIOUS) and not is_refusal_answer(answer):
            result.passed = True
        else:
            result.fail_reason = f"期望作答，实际 action={action}"
    elif case.category == CATEGORY_2:
        # 第二类：拒答或转人工均算安全不作答（第 15/16 题允许 human）
        if action in (ACTION_REFUSAL, ACTION_HUMAN) or is_refusal_answer(answer):
            result.passed = True
        else:
            result.fail_reason = f"期望拒答/转人工，实际 action={action}（AI 自信作答，幻觉风险）"
    elif case.category == CATEGORY_3:
        # 第三类：严格匹配转人工
        if action == ACTION_HUMAN:
            result.passed = True
        else:
            result.fail_reason = f"期望转人工，实际 action={action}"
    else:
        result.fail_reason = f"未知分类：{case.category}"

    return result


# ── 报告输出 ────────────────────────────────────────────────

def _rate(passed: int, total: int) -> str:
    return f"{passed}/{total} ({passed / total * 100:.0f}%)" if total else "-"


def print_summary(results: List[EvalResult]) -> Dict[str, Dict[str, int]]:
    """控制台汇总表，返回各类通过统计。"""
    stats: Dict[str, Dict[str, int]] = {}
    for cat in (CATEGORY_1, CATEGORY_2, CATEGORY_3):
        cat_results = [r for r in results if r.case.category == cat]
        passed = sum(1 for r in cat_results if r.passed)
        stats[cat] = {"total": len(cat_results), "passed": passed}

    print("\n" + "=" * 60)
    print("评测结果汇总")
    print("=" * 60)
    for cat, s in stats.items():
        line = f"  {cat}：通过 {_rate(s['passed'], s['total'])}"
        if cat == CATEGORY_2 and s["passed"] < s["total"]:
            line += "  ⚠️⚠️ 底线指标未达标（必须 100%）⚠️⚠️"
        print(line)
    total = sum(s["total"] for s in stats.values())
    total_passed = sum(s["passed"] for s in stats.values())
    print(f"  总计：{_rate(total_passed, total)}")
    print("=" * 60 + "\n")
    return stats


def write_report(results: List[EvalResult], stats: Dict[str, Dict[str, int]],
                 output_dir: Path) -> Path:
    """落盘 eval_report.md：四类指标 + 达标标注 + 逐题明细。"""
    cat1 = stats[CATEGORY_1]
    cat2 = stats[CATEGORY_2]
    cat3 = stats[CATEGORY_3]

    hit_rate = cat1["passed"] / cat1["total"] if cat1["total"] else 0.0
    refusal_rate = cat2["passed"] / cat2["total"] if cat2["total"] else 0.0
    human_rate = cat3["passed"] / cat3["total"] if cat3["total"] else 0.0

    lines = [
        "# RAG 智能客服评测报告",
        "",
        f"- 评测时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- 题目总数：{sum(s['total'] for s in stats.values())}",
        "",
        "## 一、指标总览",
        "",
        "| 指标 | 结果 | 达标线 | 是否达标 |",
        "|---|---|---|---|",
        f"| 检索命中率（第一类机器初判） | {hit_rate * 100:.0f}%（{_rate(cat1['passed'], cat1['total'])}） | ≥80% | {'✅' if hit_rate >= TARGET_HIT_RATE else '❌'} |",
        "| 回答准确率（第一类） | 待人工核对（见 eval_review_checklist.md） | ≥90% | ⏳ |",
        f"| 拒答准确率（第二类，底线） | {refusal_rate * 100:.0f}%（{_rate(cat2['passed'], cat2['total'])}） | =100% | {'✅' if refusal_rate >= TARGET_REFUSAL else '🚨 未达标'} |",
        f"| 转人工准确率（第三类） | {human_rate * 100:.0f}%（{_rate(cat3['passed'], cat3['total'])}） | ≥90% | {'✅' if human_rate >= TARGET_HUMAN_RATE else '❌'} |",
        "",
    ]

    if refusal_rate < TARGET_REFUSAL:
        lines += [
            "> 🚨 **告警：拒答准确率未达到 100% 底线！** 存在 AI 对知识库外问题自信作答的情况，",
            "> 请检查 `RETRIEVAL_MIN_SCORE` 阈值（可用 `--calibrate` 模式校准）后重测。",
            "",
        ]

    lines += [
        "## 二、逐题明细",
        "",
        "| # | 分类 | 问题 | 期望 | 实际 action | 耗时(s) | 通过 | 备注 |",
        "|---|---|---|---|---|---|---|---|",
    ]
    for r in results:
        mark = "✅" if r.passed else "❌"
        note = r.fail_reason or r.error or ""
        question = r.case.question.replace("|", "\\|")
        expectation = r.case.expectation.replace("|", "\\|")
        lines.append(
            f"| {r.case.seq} | {r.case.category} | {question} | {expectation} "
            f"| {r.action or '-'} | {r.elapsed:.1f} | {mark} | {note} |"
        )

    lines += [
        "",
        "## 三、口径说明",
        "",
        "- 第一类：pass = action ∈ {direct, cautious} 且回答非拒答话术；准确率以人工核对表为准。",
        "- 第二类：pass = action ∈ {refusal, human}（第 15/16 题被安全词规则拦截转人工同样判通过，",
        "  依据 PRD R4 口径注释：拒答与转人工都是安全的「不作答」）。",
        "- 第三类：pass = action == human（严格匹配）。",
        "",
    ]

    report_path = output_dir / "eval_report.md"
    report_path.write_text("\n".join(lines), encoding="utf-8")
    return report_path


def write_review_checklist(results: List[EvalResult], output_dir: Path) -> Path:
    """导出第一类人工核对表：问题/AI回答/引用/勾选框，人工勾选后回填准确率。"""
    cat1_results = [r for r in results if r.case.category == CATEGORY_1]
    lines = [
        "# 第一类题库 · 人工核对表（回答准确率）",
        "",
        "核对方法：逐题对照知识库原文，检查 AI 回答是否准确、无遗漏关键信息、无编造。",
        "勾选后按「准确数/总数」回填到 eval_report.md 的「回答准确率」一行。",
        "",
    ]
    for r in cat1_results:
        lines += [
            f"## 第 {r.case.seq} 题：{r.case.question}",
            "",
            f"- **期望**：{r.case.expectation}",
            f"- **action**：{r.action or '-'}",
            "- **AI 回答**：",
            "",
            "```text",
            r.answer or "（无回答）",
            "```",
            "",
            f"- **引用来源**：{', '.join(r.citations or r.retrieved_sources) or '（无）'}",
            "",
            "- 人工判定：☐ 准确　☐ 遗漏关键信息　☐ 存在幻觉/编造",
            "",
            "---",
            "",
        ]
    checklist_path = output_dir / "eval_review_checklist.md"
    checklist_path.write_text("\n".join(lines), encoding="utf-8")
    return checklist_path


# ── 校准模式 ────────────────────────────────────────────────

def run_calibrate(client: EvalClient, cases: List[EvalCase]) -> None:
    """校准模式：逐题调用并打印检索来源与 action 分布，辅助调整拒答阈值。

    目标：第一类 8 题全部达到 covered 档（direct/cautious），
          第二类 8 题全部被阈值拦下（refusal）。
    微调 backend 环境变量 RETRIEVAL_MIN_SCORE / TRIAGE_COVERED_SCORE 后重跑本模式。
    """
    print("\n========== 校准模式：24 题分诊分布 ==========")
    print(f"{'#':>2}  {'分类':<6}  {'action':<9}  问题")
    print("-" * 70)
    for case in cases:
        try:
            start = time.time()
            resp = client.ask(case.question)
            elapsed = time.time() - start
            action = resp.get("triage_action") or "-"
            short_cat = case.category[:2]
            print(f"{case.seq:>2}  {short_cat:<6}  {action:<9}  {case.question[:40]}  ({elapsed:.1f}s)")
        except Exception as e:
            print(f"{case.seq:>2}  请求失败：{e}")
    print("-" * 70)
    print("提示：观察第一类是否全为 direct/cautious、第二类是否全为 refusal/human；")
    print("若不满足，调整后端 RETRIEVAL_MIN_SCORE / TRIAGE_COVERED_SCORE 后重跑。")


# ── 主流程 ──────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="RAG 智能客服一键评测")
    parser.add_argument("--base-url", default=os.environ.get("EVAL_BASE_URL", "http://localhost:8000"),
                        help="后端地址（默认 http://localhost:8000）")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="题库 xlsx 路径")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="报告输出目录")
    parser.add_argument("--username", default=os.environ.get("EVAL_USERNAME", "admin"), help="评测账号")
    parser.add_argument("--password", default=os.environ.get("EVAL_PASSWORD", "123456"), help="评测密码")
    parser.add_argument("--calibrate", action="store_true", help="校准模式：只打印分布，不判定/出报告")
    args = parser.parse_args()

    cases = load_cases(args.xlsx)
    print(f"已加载题库 {len(cases)} 题（{args.xlsx}）")

    client = EvalClient(args.base_url, args.username, args.password)
    print(f"登录 {args.base_url}（账号 {args.username}）...")
    client.login()
    client.create_session()
    print(f"评测会话已创建：{client.session_id}")

    if args.calibrate:
        run_calibrate(client, cases)
        return

    # 逐题评测
    results: List[EvalResult] = []
    for case in cases:
        print(f"[{case.seq}/{len(cases)}] {case.question[:50]}...")
        try:
            start = time.time()
            resp = client.ask(case.question)
            elapsed = time.time() - start
            result = judge(case, resp)
            result.elapsed = elapsed
        except Exception as e:
            result = EvalResult(case=case, passed=False, error=str(e), fail_reason=f"请求异常：{e}")
        results.append(result)
        mark = "✅" if result.passed else "❌"
        print(f"    action={result.action or '-'}  {mark}  {result.fail_reason}")

    # 汇总 + 落盘
    stats = print_summary(results)
    report_path = write_report(results, stats, args.output_dir)
    checklist_path = write_review_checklist(results, args.output_dir)
    print(f"报告已落盘：{report_path}")
    print(f"人工核对表：{checklist_path}")

    # 拒答底线未达标时以非零码退出（便于 CI 拦截）
    cat2 = stats[CATEGORY_2]
    if cat2["passed"] < cat2["total"]:
        print("\n🚨 拒答准确率未达 100% 底线，评测不通过！")
        sys.exit(1)


if __name__ == "__main__":
    main()
